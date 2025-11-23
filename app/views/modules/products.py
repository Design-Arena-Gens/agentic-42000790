from __future__ import annotations

from PySide6.QtCore import Slot
from PySide6.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QLineEdit, QPushButton, QTableWidget, QTableWidgetItem, QSpinBox, QMessageBox, QFileDialog

from app.core.db import Database
from openpyxl import Workbook, load_workbook
import csv
from pathlib import Path


class ProductsView(QWidget):
    def __init__(self, db: Database, parent=None):
        super().__init__(parent)
        self.db = db
        self._page = 1
        self._page_size = 20
        self._build_ui()
        self._load()

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)
        top = QHBoxLayout()
        self.searchEdit = QLineEdit(self)
        self.searchEdit.setPlaceholderText(self.tr("Rechercher produit..."))
        self.addBtn = QPushButton(self.tr("Ajouter"))
        self.editBtn = QPushButton(self.tr("Modifier"))
        self.delBtn = QPushButton(self.tr("Supprimer"))
        self.exportCsvBtn = QPushButton(self.tr("Export CSV"))
        self.exportXlsxBtn = QPushButton(self.tr("Export Excel"))
        self.importCsvBtn = QPushButton(self.tr("Import CSV"))
        self.importXlsxBtn = QPushButton(self.tr("Import Excel"))
        for w in [self.searchEdit, self.addBtn, self.editBtn, self.delBtn, self.exportCsvBtn, self.exportXlsxBtn, self.importCsvBtn, self.importXlsxBtn]:
            top.addWidget(w)
        layout.addLayout(top)

        self.table = QTableWidget(self)
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels([self.tr("ID"), self.tr("SKU"), self.tr("Nom (FR)"), self.tr("Nom (AR)"), self.tr("Unit"), self.tr("Prix HT")])
        self.table.setSelectionBehavior(self.table.SelectRows)
        self.table.setEditTriggers(self.table.NoEditTriggers)
        layout.addWidget(self.table)

        bottom = QHBoxLayout()
        self.pageSpin = QSpinBox(self)
        self.pageSpin.setMinimum(1)
        self.pageSpin.setValue(1)
        self.pageSizeSpin = QSpinBox(self)
        self.pageSizeSpin.setRange(10, 200)
        self.pageSizeSpin.setValue(self._page_size)
        self.refreshBtn = QPushButton(self.tr("Actualiser"))
        bottom.addWidget(self.pageSpin)
        bottom.addWidget(self.pageSizeSpin)
        bottom.addWidget(self.refreshBtn)
        layout.addLayout(bottom)

        self.searchEdit.textChanged.connect(self._load)
        self.refreshBtn.clicked.connect(self._load)
        self.pageSpin.valueChanged.connect(self._on_page_change)
        self.pageSizeSpin.valueChanged.connect(self._on_page_size_change)
        self.addBtn.clicked.connect(self._add)
        self.editBtn.clicked.connect(self._edit)
        self.delBtn.clicked.connect(self._delete)
        self.exportCsvBtn.clicked.connect(self._export_csv)
        self.exportXlsxBtn.clicked.connect(self._export_xlsx)
        self.importCsvBtn.clicked.connect(self._import_csv)
        self.importXlsxBtn.clicked.connect(self._import_xlsx)

    def _query(self):
        base = "SELECT id, sku, name_fr, name_ar, unit, price_ht FROM products"
        params = []
        q = self.searchEdit.text().strip()
        if q:
            base += " WHERE (sku LIKE ? OR name_fr LIKE ? OR name_ar LIKE ?)"
            like = f\"%{q}%\"
            params.extend([like, like, like])
        base += " ORDER BY id DESC LIMIT ? OFFSET ?"
        params.extend([self._page_size, (self._page - 1) * self._page_size])
        return base, params

    def _load(self) -> None:
        sql, params = self._query()
        rows = self.db.query(sql, params)
        self.table.setRowCount(len(rows))
        for r, row in enumerate(rows):
            self.table.setItem(r, 0, QTableWidgetItem(str(row["id"])))
            self.table.setItem(r, 1, QTableWidgetItem(row["sku"]))
            self.table.setItem(r, 2, QTableWidgetItem(row["name_fr"]))
            self.table.setItem(r, 3, QTableWidgetItem(row["name_ar"]))
            self.table.setItem(r, 4, QTableWidgetItem(row["unit"]))
            self.table.setItem(r, 5, QTableWidgetItem(str(row["price_ht"])))
        self.table.resizeColumnsToContents()

    @Slot()
    def _on_page_change(self, val: int) -> None:
        self._page = max(1, val)
        self._load()

    @Slot()
    def _on_page_size_change(self, val: int) -> None:
        self._page_size = val
        self._load()

    def _current_id(self) -> int | None:
        indexes = self.table.selectionModel().selectedRows()
        if not indexes:
            return None
        return int(self.table.item(indexes[0].row(), 0).text())

    @Slot()
    def _add(self) -> None:
        self.db.execute("INSERT INTO products (sku, name_fr, name_ar, unit, price_ht) VALUES (?, ?, ?, ?, ?);", ("SKU", "Produit", "????", "u", 100.0))
        self._load()

    @Slot()
    def _edit(self) -> None:
        pid = self._current_id()
        if not pid:
            return
        self.db.execute("UPDATE products SET name_fr=name_fr||' *' WHERE id=?;", (pid,))
        self._load()

    @Slot()
    def _delete(self) -> None:
        pid = self._current_id()
        if not pid:
            return
        if QMessageBox.question(self, self.tr("Confirmer"), self.tr("Supprimer ce produit ?")) != QMessageBox.Yes:
            return
        self.db.execute("DELETE FROM products WHERE id=?;", (pid,))
        self._load()

    @Slot()
    def _export_csv(self) -> None:
        path, _ = QFileDialog.getSaveFileName(self, self.tr("Exporter CSV"), "produits.csv", self.tr("CSV (*.csv)"))
        if not path:
            return
        rows = self.db.query("SELECT sku, name_fr, name_ar, unit, price_ht FROM products ORDER BY id;")
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["sku", "name_fr", "name_ar", "unit", "price_ht"])
            for r in rows:
                writer.writerow([r["sku"], r["name_fr"], r["name_ar"], r["unit"], r["price_ht"]])

    @Slot()
    def _export_xlsx(self) -> None:
        path, _ = QFileDialog.getSaveFileName(self, self.tr("Exporter Excel"), "produits.xlsx", self.tr("Excel (*.xlsx)"))
        if not path:
            return
        wb = Workbook()
        ws = wb.active
        ws.append(["sku", "name_fr", "name_ar", "unit", "price_ht"])
        rows = self.db.query("SELECT sku, name_fr, name_ar, unit, price_ht FROM products ORDER BY id;")
        for r in rows:
            ws.append([r["sku"], r["name_fr"], r["name_ar"], r["unit"], r["price_ht"]])
        wb.save(path)

    @Slot()
    def _import_csv(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, self.tr("Importer CSV"), "", self.tr("CSV (*.csv)"))
        if not path:
            return
        with open(path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                self.db.execute(
                    "INSERT OR REPLACE INTO products (sku, name_fr, name_ar, unit, price_ht) VALUES (?, ?, ?, ?, ?);",
                    (row.get("sku",""), row.get("name_fr",""), row.get("name_ar",""), row.get("unit","u"), float(row.get("price_ht", "0") or 0)),
                )
        self._load()

    @Slot()
    def _import_xlsx(self) -> None:
        path, _ = QFileDialog.getOpenFileName(self, self.tr("Importer Excel"), "", self.tr("Excel (*.xlsx)"))
        if not path:
            return
        wb = load_workbook(path)
        ws = wb.active
        first = True
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue
            sku, name_fr, name_ar, unit, price_ht = row
            self.db.execute(
                "INSERT OR REPLACE INTO products (sku, name_fr, name_ar, unit, price_ht) VALUES (?, ?, ?, ?, ?);",
                (sku or "", name_fr or "", name_ar or "", unit or "u", float(price_ht or 0)),
            )
        self._load()

